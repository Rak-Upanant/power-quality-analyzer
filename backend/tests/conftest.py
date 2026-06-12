# backend/tests/conftest.py
"""
Shared pytest fixtures.

The real CA8335 dummy workbooks are not committed to the repo, so these
fixtures BUILD tiny synthetic workbooks in memory that reproduce the exact
sheet layout of each meter format (header rows, preamble rows, date formats).
This keeps the test suite self-contained and runnable anywhere, including CI.

Two formats are produced:
  * SN210210 — native schema (`Trend`, `Vh/Ah Harmonic %`), Thai Buddhist dates
  * SN3007   — legacy schema (`Recording` + `Energy`, `V/A H Harmonic %`),
               Gregorian M/D/YYYY 12-hour dates
"""
import io

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Low-level workbook helper
# ---------------------------------------------------------------------------

def _write_rows(ws, rows, start_row=1):
    """Write a list-of-lists into a worksheet starting at `start_row` (1-based)."""
    for r, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)


def _book_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# SN210210 (native) builder
# ---------------------------------------------------------------------------

# Columns the analyzer reads from the Trend sheet (subset is fine — missing
# columns default to 0 downstream).
_SN210_TREND_COLS = [
    "Date", "Time", "Frequency",
    "U1 RMS", "U2 RMS", "U3 RMS",
    "V1 RMS", "V2 RMS", "V3 RMS",
    "A1 RMS", "A2 RMS", "A3 RMS",
    "V1 THD", "V2 THD", "V3 THD",
    "U1 THD", "U2 THD", "U3 THD",
    "A1 THD", "A2 THD", "A3 THD",
    "W Total", "var Total", "VA Total",
    "Wh Total", "varh Total", "VAh Total",
    "PF1", "PF2", "PF3", "PF Mean",
    "Vunb", "Aunb",
]


def _sn210_trend_data_row(i):
    """One Trend data row. Thai Buddhist date 31/3/2569 == 2026-03-31."""
    minute = 15 + i * 5
    hh = 14 + minute // 60
    mm = minute % 60
    time = f"{hh:02d}:{mm:02d}:00"
    wh = 1000 + i * 50          # running counter that starts non-zero
    return [
        "31/3/2569", time, 50.0,
        400.0, 401.0, 402.0,
        231.0, 231.5, 232.0,
        100.0 + i, 101.0 + i, 99.0 + i,
        2.5, 2.6, 2.4,          # V THD
        2.0, 2.1, 1.9,          # U THD
        8.0, 8.5, 7.5,          # A THD
        70000.0, 12000.0, 71000.0,
        float(wh), float(wh) // 3, float(wh) + 100,
        0.92, 0.93, 0.91, 0.92,
        0.8, 1.2,
    ]


def build_sn210210_bytes(n_rows=12) -> bytes:
    wb = openpyxl.Workbook()

    # ── Trend (header on row 7, units row 8, blank row 9, data from row 10) ──
    tr = wb.active
    tr.title = "Trend"
    _write_rows(tr, [
        ["C.A 8335", "Serial number", "210210"],
        ["Trend", "MCC1"],
        ["Date Started", "Time Started", "Date Ended", "Time Ended"],
        ["31/3/2569", "14:15:00", "7/4/2569", "14:15:00"],
        ["Connection Type: 3-Phase 4-Wire"],
        [],
        _SN210_TREND_COLS,                                  # row 7 — header
        ["", "", "Hz"] + ["V"] * 9 + ["%"] * 9 + ["W", "var", "VA", "Wh", "varh", "VAh"] + ["", "", "", "", "%", "%"],  # row 8 units
        [],                                                  # row 9 blank
    ])
    _write_rows(tr, [_sn210_trend_data_row(i) for i in range(n_rows)], start_row=10)

    # ── Harmonic sheets (header row 1, units row 2, data row 3+) ────────────
    for sheet, prefix in (("Vh Harmonic %", "V"), ("Ah Harmonic %", "A")):
        ws = wb.create_sheet(sheet)
        cols = ["Date", "Time"]
        for h in (3, 5, 7):
            for p in (1, 2, 3):
                cols.append(f"{prefix}{p}h{h}")
        rows = [cols, ["", ""] + ["%"] * (len(cols) - 2)]
        for i in range(n_rows):
            minute = 15 + i * 5
            hh = 14 + minute // 60
            time = f"{hh:02d}:{minute % 60:02d}:00"
            rows.append(["31/3/2569", time] + [1.5, 1.4, 1.3, 3.0, 2.9, 2.8, 0.9, 0.8, 0.7])
        _write_rows(ws, rows)

    return _book_to_bytes(wb)


# ---------------------------------------------------------------------------
# SN3007 (legacy) builder
# ---------------------------------------------------------------------------

# SN3007 Recording columns we care about (the adapter maps these).
_SN3007_REC_COLS = [
    "Date:", "Time:", "F",
    "U12 RMS", "U23 RMS", "U31 RMS",
    "V1 RMS", "V2 RMS", "V3 RMS",
    "A1 RMS", "A2 RMS", "A3 RMS",
    "V1 THDf", "V2 THDf", "V3 THDf",
    "U12 THDf", "U23 THDf", "U31 THDf",
    "A1 THDf", "A2 THDf", "A3 THDf",
    "Vunb (u2)", "Aunb (u2)",
    "FK1", "FK2", "FK3",
    "P1 (W)", "P2 (W)", "P3 (W)", "PT (W)",
    "S1 (VA)", "S2 (VA)", "S3 (VA)", "ST (VA)",
    "PF1", "PF2", "PF3", "PFT",
    "Cos φ1 (DPF)", "Cos φ2 (DPF)", "Cos φ3 (DPF)", "Cos φT (DPF)",
]


def _sn3007_rec_data_row(i):
    """Gregorian date 4/30/2026, 12-hour clock."""
    minute = 10 + i
    time = f"11:{minute % 60:02d}:00 AM"
    return [
        "4/30/2026", time, 50.0,
        400.0, 401.0, 402.0,
        231.0, 231.5, 232.0,
        100.0 + i, 101.0 + i, 99.0 + i,
        2.5, 2.6, 2.4,
        2.0, 2.1, 1.9,
        8.0, 8.5, 7.5,
        0.8, 1.2,
        1.05, 1.06, 1.04,
        23000.0, 23500.0, 23000.0, 69500.0,
        24000.0, 24500.0, 24000.0, 72500.0,
        0.92, 0.93, 0.91, 0.92,
        0.95, 0.96, 0.94, 0.95,
    ]


_SN3007_ENERGY_COLS = [
    "Date:", "Time:",
    "Ep1 (Wh)", "Ep2 (Wh)", "Ep3 (Wh)", "EpT (Wh)",
    "En1 (varh)", "En2 (varh)", "En3 (varh)", "EnT (varh)",
    "Es1 (VAh)", "Es2 (VAh)", "Es3 (VAh)", "EsT (VAh)",
]


def _sn3007_energy_data_row(i):
    minute = 10 + i
    time = f"11:{minute % 60:02d}:00 AM"
    wh = i * 150.0
    return [
        "4/30/2026", time,
        wh, wh, wh, wh * 3,
        wh / 3, wh / 3, wh / 3, wh,
        wh, wh, wh, wh * 3,
    ]


def build_sn3007_bytes(n_rows=12) -> bytes:
    wb = openpyxl.Workbook()

    # ── Recording (title row 1, header row 2, units row 3, blank row 4) ─────
    rec = wb.active
    rec.title = "Recording"
    _write_rows(rec, [
        ["Recording"],
        _SN3007_REC_COLS,                                   # row 2 — header
        ["", "", "Hz"] + [""] * (len(_SN3007_REC_COLS) - 3),  # row 3 units
        [],                                                  # row 4 blank
    ])
    _write_rows(rec, [_sn3007_rec_data_row(i) for i in range(n_rows)], start_row=5)

    # ── Energy (same layout as Recording) ───────────────────────────────────
    en = wb.create_sheet("Energy")
    _write_rows(en, [
        ["Energy"],
        _SN3007_ENERGY_COLS,
        ["", ""] + ["Wh"] * (len(_SN3007_ENERGY_COLS) - 2),
        [],
    ])
    _write_rows(en, [_sn3007_energy_data_row(i) for i in range(n_rows)], start_row=5)

    # ── Harmonic sheets (title row 1, header row 2, units row 3) ────────────
    for sheet, prefix in (("V H Harmonic %", "V"), ("A H Harmonic %", "A")):
        ws = wb.create_sheet(sheet)
        cols = ["Date:", "Time:"]
        for h in (3, 5, 7):
            for p in (1, 2, 3):
                cols.append(f"{prefix}{p} H{h}")
        rows = [[sheet], cols, ["", ""] + ["% f"] * (len(cols) - 2)]
        for i in range(n_rows):
            minute = 10 + i
            time = f"11:{minute % 60:02d}:00 AM"
            # '- - -' placeholder must survive → NaN downstream
            vals = [1.5, 1.4, 1.3, 3.0, 2.9, 2.8, "- - -", "- - -", "- - -"]
            rows.append(["4/30/2026", time] + vals)
        _write_rows(ws, rows)

    return _book_to_bytes(wb)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sn210210_bytes() -> bytes:
    return build_sn210210_bytes()


@pytest.fixture
def sn3007_bytes() -> bytes:
    return build_sn3007_bytes()
